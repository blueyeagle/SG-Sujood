#!/usr/bin/env python3
"""Publish moderator-approved submissions into the live directory (spaces.json).

Flow: a moderator sets a row's Status to "Approved" in "Prayer Space for Review.xlsx" and
commits. This script geocodes the building via OneMap, appends a SpaceRecord to spaces.json,
and flips the row's Status to "Published" (so it isn't re-added). Rows that can't be geocoded
are marked "Needs address"; exact-name duplicates are marked "Duplicate".
"""
import json
import os
import re
import time
import urllib.parse
import urllib.request

import openpyxl

WB = "Prayer Space for Review.xlsx"
SPACES = "spaces.json"


def slug(s):
    return re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")


def onemap(query):
    if not query:
        return None
    url = "https://www.onemap.gov.sg/api/common/elastic/search?" + urllib.parse.urlencode(
        {"searchVal": query, "returnGeom": "Y", "getAddrDetails": "Y", "pageNum": 1}
    )
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "waqtsg-publish"})
        d = json.load(urllib.request.urlopen(req, timeout=25))
        if d.get("found") and d.get("results"):
            r = d["results"][0]
            postal = r.get("POSTAL", "")
            return {
                "lat": round(float(r["LATITUDE"]), 6),
                "lng": round(float(r["LONGITUDE"]), 6),
                "address": r.get("ADDRESS", "") or None,
                "postal": (postal if postal and postal != "NIL" else None),
            }
    except Exception:
        pass
    return None


def region_for(lat, lng):
    if lat >= 1.38:
        return "North"
    if lng >= 103.90:
        return "East"
    if lng <= 103.78:
        return "West"
    return "Central"


data = json.load(open(SPACES))
existing_ids = {s["id"] for s in data["spaces"]}
existing_names = {s["name"].strip().lower() for s in data["spaces"]}

wb = openpyxl.load_workbook(WB)
ws = wb["Submissions"] if "Submissions" in wb.sheetnames else wb.active
headers = [c.value for c in ws[1]]

def idx(name):
    return headers.index(name)

i_building = idx("Building / Mall")
i_floor = idx("Floor & Landmark")
i_type = idx("Type")
i_status = idx("Status")

added = 0
for row in ws.iter_rows(min_row=2):
    status = (row[i_status].value or "").strip().lower()
    if status != "approved":
        continue
    name = (row[i_building].value or "").strip()
    if not name:
        continue
    if name.lower() in existing_names:
        row[i_status].value = "Duplicate"
        continue
    geo = onemap(name)
    if not geo:
        row[i_status].value = "Needs address"
        continue

    stype = (row[i_type].value or "Musollah").strip()
    category = "masjid" if stype.lower() == "masjid" else "musollah"
    base = slug(name) or "space"
    sid = "u-" + base
    n = 2
    while sid in existing_ids:
        sid = f"u-{base}-{n}"
        n += 1

    rec = {
        "id": sid, "name": name, "category": category, "type": stype,
        "regionGroup": region_for(geo["lat"], geo["lng"]), "area": None,
        "address": geo["address"], "postal": geo["postal"],
        "floor": (row[i_floor].value or None), "genderSegregated": None,
        "facilities": None, "access": "Public", "capacity": None,
        "yearEst": None, "heritage": None,
        "notes": "Community submitted via Waqt SG.",
        "lat": geo["lat"], "lng": geo["lng"],
    }
    data["spaces"].append(rec)
    existing_ids.add(sid)
    existing_names.add(name.lower())
    row[i_status].value = "Published"
    added += 1
    time.sleep(0.3)

if added:
    data["count"] = len(data["spaces"])
    json.dump(data, open(SPACES, "w"), ensure_ascii=False, indent=1)
    wb.save(WB)

print(f"published {added} space(s)")
with open(os.environ.get("GITHUB_OUTPUT", os.devnull), "a") as f:
    f.write(f"added={added}\n")
