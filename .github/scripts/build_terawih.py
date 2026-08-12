#!/usr/bin/env python3
"""Build terawih.json from a terawih-venues workbook.

Reads "Terawih Venues.xlsx" (or the path in argv[1]), geocodes each venue via OneMap, and
writes terawih.json. Flexible about columns — it looks (case-insensitively) for:
  • name:    "Name" / "Mosque" / "Masjid" / "Venue" / "Prayer space"
  • address: "Address" / "Postal" / "Location"           (used for geocoding + display)
  • note:    "Note" / "Remarks" / "Details"              (optional)
No rakaat counts or start times are stored.
"""
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request

import openpyxl

WORKBOOK = sys.argv[1] if len(sys.argv) > 1 else "Terawih Venues.xlsx"
OUT = "terawih.json"

NAME_KEYS = ["name", "mosque", "masjid", "venue", "prayer space", "surau"]
ADDR_KEYS = ["address", "postal", "location"]
NOTE_KEYS = ["note", "remarks", "details"]


def slug(s):
    return re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")


def onemap(query, tries=2):
    if not query:
        return None
    url = "https://www.onemap.gov.sg/api/common/elastic/search?" + urllib.parse.urlencode(
        {"searchVal": query, "returnGeom": "Y", "getAddrDetails": "Y", "pageNum": 1}
    )
    for _ in range(tries):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "sgsujood-terawih"})
            d = json.load(urllib.request.urlopen(req, timeout=25))
            if d.get("found") and d.get("results"):
                r = d["results"][0]
                return round(float(r["LATITUDE"]), 6), round(float(r["LONGITUDE"]), 6)
            return None
        except Exception:
            time.sleep(0.8)
    return None


def find_header(ws):
    # A header row has >=2 labelled columns, one matching a name key and one an address key.
    def scan(require_addr):
        for row in ws.iter_rows(min_row=1, max_row=20):
            labels = [str(c.value).strip().lower() if c.value else "" for c in row]
            if len([l for l in labels if l]) < 2:
                continue                      # skip single-cell title rows
            if col_for(labels, NAME_KEYS) is not None and (not require_addr or col_for(labels, ADDR_KEYS) is not None):
                return row[0].row, labels
        return None, None
    r = scan(True)
    return r if r[0] is not None else scan(False)


def col_for(labels, keys):
    for i, l in enumerate(labels):
        if any(k == l or k in l for k in keys):
            return i
    return None


wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
ws = wb.active
header_row, labels = find_header(ws)
if header_row is None:
    print("No header row found; writing empty venues.")
    venues = []
else:
    i_name = col_for(labels, NAME_KEYS)
    i_addr = col_for(labels, ADDR_KEYS)
    i_note = col_for(labels, NOTE_KEYS)
    venues = []
    seen = set()
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        cells = [c for c in row if c is not None and str(c).strip() != ""]
        if not cells:
            continue
        name = (str(row[i_name]).strip() if i_name is not None and row[i_name] else "")
        if not name or len(cells) == 1:   # skip blanks / section headers
            continue
        addr = (str(row[i_addr]).strip() if i_addr is not None and row[i_addr] else None)
        note = (str(row[i_note]).strip() if i_note is not None and row[i_note] else None)
        postal = None
        if addr:
            m = re.search(r"\b[Ss]?(\d{6})\b", addr)
            postal = m.group(1) if m else None
        coord = onemap(postal) or onemap(name + " Singapore") or onemap(addr)
        vid = slug(name) or "venue"
        base = vid
        n = 2
        while vid in seen:
            vid = f"{base}-{n}"; n += 1
        seen.add(vid)
        venues.append({
            "id": vid, "name": name, "address": addr, "note": note,
            "lat": (coord[0] if coord else None), "lng": (coord[1] if coord else None),
        })
        time.sleep(0.3)

geocoded = sum(1 for v in venues if v["lat"])
print(f"parsed {len(venues)} venue(s), geocoded {geocoded}")

out = {
    "_note": "Terawih venues, regenerated from the terawih workbook. No rakaat/start-time.",
    "source": os.path.basename(WORKBOOK),
    "updated": os.environ.get("BUILD_DATE", ""),
    "venues": venues,
}
json.dump(out, open(OUT, "w"), ensure_ascii=False, indent=1)
print("wrote", OUT)
